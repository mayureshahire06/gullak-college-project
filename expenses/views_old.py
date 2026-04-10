import calendar
import csv
import json
import traceback
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl
from allauth.account.models import EmailAddress
from allauth.socialaccount.models import SocialAccount
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.management import call_command
from django.db import IntegrityError
from django.db.models import Count, Q, Sum
from django.db.models.functions import ExtractWeekDay, TruncDay, TruncMonth
from django.forms import modelformset_factory
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.formats import date_format
from django.utils.html import escape, format_html, format_html_join, mark_safe
from django.utils.translation import gettext as _
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import (
    CreateView,
    DeleteView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

from finance_tracker.ai_utils import predict_category_ai

from .forms import (
    CategoryForm,
    ContactForm,
    CustomSignupForm,
    ExpenseForm,
    GoalContributionForm,
    IncomeForm,
    LanguageUpdateForm,
    ProfileUpdateForm,
    RecurringTransactionForm,
    SavingsGoalForm,
)
from .models import (
    CURRENCY_CHOICES,
    Category,
    Expense,
    Income,
    Notification,
    RecurringTransaction,
    SavingsGoal,
    SubscriptionPlan,
    UserProfile,
)
from .utils import BOOTSTRAP_ICONS, generate_year_in_review_data, get_exchange_rate


def create_category_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': _('Category name cannot be empty.')}, status=400)
            
            # Check Limits
            current_count = Category.objects.filter(user=request.user).count()
            limit = 5 # Free
            if request.user.profile.is_plus:
                limit = 10
            if request.user.profile.is_pro:
                limit = float('inf')

            if current_count >= limit:
                 return JsonResponse({
                     'success': False, 
                     'error': _('Category limit reached for your plan (%(limit)s). Upgrade to Pro to unlock unlimited categories, AI insights, and more.') % {'limit': limit}
                 }, status=403)

            category = Category.objects.create(user=request.user, name=name)
            return JsonResponse({'success': True, 'id': category.id, 'name': category.name})
            
        except IntegrityError:
            return JsonResponse({'success': False, 'error': _('This category already exists.')}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': _('Invalid request method.')}, status=405)


def resend_verification_email(request):
    """
    AJAX view to resend verification email.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            # If email is not in body, try to get from logged in user
            if not email and request.user.is_authenticated:
                email = request.user.email
            
            # Fallback: Check allauth session key
            if not email:
                email = request.session.get('account_email')

            if not email:
                return JsonResponse({'success': False, 'error': 'Email is missing.'}, status=400)
            
            try:
                # Case-insensitive lookup just in case
                email_address = EmailAddress.objects.filter(email__iexact=email).first()
                if not email_address:
                     return JsonResponse({'success': False, 'error': f'Email {email} not found in system.'}, status=404)
                
                # Check if already verified
                if email_address.verified:
                    return JsonResponse({'success': True, 'message': 'Email already verified.'})

                email_address.send_confirmation(request)
                return JsonResponse({'success': True, 'message': 'Verification email sent!'})

            except Exception as e:
                # Log the actual error for debugging
                
                print(traceback.format_exc())
                return JsonResponse({'success': False, 'error': f'Send failed: {str(e)}'}, status=500)
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Server Error: {str(e)}'}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    Ensures data is always fresh (current month).
    """
    # Clear messages
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        
        # Check if data is stale (i.e. not from this month)
        # We check the latest expense. If no expenses or old date, refresh.
        last_expense = Expense.objects.filter(user=user).order_by('-date').first()
        is_stale = False
        
        if not last_expense:
            is_stale = True
        else:
            today = date.today()
            if last_expense.date.month != today.month or last_expense.date.year != today.year:
                is_stale = True
        
        if is_stale:
            # Data is old, refresh it
            call_command('setup_demo_user')
            # Refetch the new user object since the old one might have been deleted/recreated
            user = User.objects.get(username='demo')

    except User.DoesNotExist:
        # User doesn't exist, create it
        call_command('setup_demo_user')
        user = User.objects.get(username='demo')

    # Manually set the backend to allow login without authentication
    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
    return redirect('home')

def demo_signup(request):
    """
    Logs out the demo user and redirects to the signup page.
    """
    logout(request)
    return redirect('signup')

# --------------------
# Mixins
# --------------------

class RecurringTransactionMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            process_user_recurring_transactions(request.user)
        return super().dispatch(request, *args, **kwargs)

    def process_recurring_transactions(self, user):
        # Deprecated: Use process_user_recurring_transactions instead
        process_user_recurring_transactions(user)

def process_user_recurring_transactions(user):
    today = date.today()
    profile = user.profile
    recurring_txs = RecurringTransaction.objects.filter(user=user, is_active=True).order_by('created_at')
    
    # Enforce Tier Limits for processing
    if not profile.is_pro:
        limit = 3 if profile.is_plus else 0
        recurring_txs = recurring_txs[:limit]
    
    new_expenses = []
    new_incomes = []
    updates_needed = []
    
    # Pre-fetch profile currency
    try:
        base_currency = user.profile.currency
    except UserProfile.DoesNotExist:
        return

    for rt in recurring_txs:
        if not rt.last_processed_date:
            current_date = rt.start_date
        else:
            current_date = rt.get_next_date(rt.last_processed_date, rt.frequency)

        # Check if processing is needed
        if current_date > today:
            continue

        # Currency conversion needed?
        exchange_rate = Decimal('1.0')
        if rt.currency != base_currency:
            exchange_rate = get_exchange_rate(rt.currency, base_currency)
        
        # Base amount calculation
        base_amount = (rt.amount * exchange_rate).quantize(Decimal('0.01'))

        # Process all pending occurrences for this transaction
        while current_date <= today:
            description = f"{rt.description} (Recurring)"
            
            if rt.transaction_type == 'EXPENSE':
                # Check for duplicates
                exists = Expense.objects.filter(
                    user=user,
                    date=current_date,
                    amount=rt.amount,
                    description=description,
                    currency=rt.currency
                ).exists()
                
                if not exists:
                    new_expenses.append(Expense(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        currency=rt.currency,
                        category=rt.category or 'Uncategorized',
                        description=description,
                        payment_method=rt.payment_method,
                        exchange_rate=exchange_rate,
                        base_amount=base_amount
                    ))
            else:
                exists = Income.objects.filter(
                    user=user,
                    date=current_date,
                    amount=rt.amount,
                    description=description,
                    currency=rt.currency
                ).exists()

                if not exists:
                    new_incomes.append(Income(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        currency=rt.currency,
                        source=rt.source or 'Other',
                        description=description,
                        exchange_rate=exchange_rate,
                        base_amount=base_amount
                    ))
            
            rt.last_processed_date = current_date
            current_date = rt.get_next_date(current_date, rt.frequency)
        
        updates_needed.append(rt)

    # Bulk Create
    if new_expenses:
        Expense.objects.bulk_create(new_expenses)
    if new_incomes:
        Income.objects.bulk_create(new_incomes)
        
    # Update Recurring Transactions (last_processed_date)
    if updates_needed:
        RecurringTransaction.objects.bulk_update(updates_needed, ['last_processed_date'])


# Custom signup view to log user in immediately
class SignUpView(generic.CreateView):
    form_class = CustomSignupForm
    success_url = reverse_lazy('onboarding')  # Redirect to onboarding after signup
    template_name = 'registration/signup.html'

class OnboardingView(LoginRequiredMixin, TemplateView):
    template_name = 'onboarding.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)

        # Redirect if user has already marked seen tutorial
        if request.user.profile.has_seen_tutorial:
            return redirect('home')
        
        # Also redirect if they already have BOTH income and expenses (not just one, to allow the flow to continue)
        has_income = Income.objects.filter(user=request.user).exists()
        has_expense = Expense.objects.filter(user=request.user).exists()
        if has_income and has_expense:
            return redirect('home')
            
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['currency_choices'] = CURRENCY_CHOICES
        context['language_choices'] = UserProfile.LANGUAGE_CHOICES
        return context

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            step = data.get('step')
            
            if step == 'setup':
                profile = request.user.profile
                profile.currency = data.get('currency', profile.currency)
                profile.language = data.get('language', profile.language)
                profile.save()
                return JsonResponse({'success': True})
            
            elif step == 'income':
                # Manually handle idempotency to avoid MultipleObjectsReturned if duplicates exist
                income_qs = Income.objects.filter(
                    user=request.user,
                    date=date.today(),
                    source=data.get('source', 'Initial Income')
                )
                if income_qs.exists():
                    income = income_qs.first()
                    income.amount = Decimal(data.get('amount', 0))
                    income.currency = request.user.profile.currency
                    income.save()
                else:
                    Income.objects.create(
                        user=request.user,
                        date=date.today(),
                        source=data.get('source', 'Initial Income'),
                        amount=Decimal(data.get('amount', 0)),
                        currency=request.user.profile.currency
                    )
                return JsonResponse({'success': True})
            
            elif step == 'budget':
                categories = data.get('categories', [])
                for cat_data in categories:
                    name = cat_data.get('name')
                    limit = cat_data.get('limit')
                    if name:
                        Category.objects.update_or_create(
                            user=request.user,
                            name=name,
                            defaults={'limit': Decimal(limit) if limit else None}
                        )
                return JsonResponse({'success': True})
            
            elif step == 'expense':
                # Manually handle idempotency to avoid MultipleObjectsReturned if duplicates exist
                expense_qs = Expense.objects.filter(
                    user=request.user,
                    date=date.today(),
                    description=data.get('description', 'Initial Expense'),
                    category=data.get('category', 'Miscellaneous')
                )
                if expense_qs.exists():
                    expense = expense_qs.first()
                    expense.amount = Decimal(data.get('amount', 0))
                    expense.currency = request.user.profile.currency
                    expense.save()
                else:
                    Expense.objects.create(
                        user=request.user,
                        date=date.today(),
                        description=data.get('description', 'Initial Expense'),
                        category=data.get('category', 'Miscellaneous'),
                        amount=Decimal(data.get('amount', 0)),
                        currency=request.user.profile.currency
                    )
                # Mark onboarding as complete
                profile = request.user.profile
                profile.has_seen_tutorial = True
                profile.save()
                return JsonResponse({'success': True})

            elif step == 'skip':
                profile = request.user.profile
                profile.has_seen_tutorial = True
                profile.save()
                return JsonResponse({'success': True})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        
        return JsonResponse({'success': False, 'error': 'Invalid step'}, status=400)

class LandingPageView(TemplateView):
    template_name = 'landing.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans_monthly'] = {p.tier: p for p in plans.filter(duration='MONTHLY')}
        context['plans_yearly'] = {p.tier: p for p in plans.filter(duration='YEARLY')}
        context['plans'] = context['plans_yearly']
        return context

class SettingsHomeView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/settings_home.html'

@login_required
def home_view(request):
    """
    Dashboard view with filters and multiple charts.
    """
    # Defensive check: Redirect to onboarding if user has NO data AND hasn't finished the flow
    try:
        if not request.user.profile.has_seen_tutorial:
            has_any_data = Expense.objects.filter(user=request.user).exists() or Income.objects.filter(user=request.user).exists()
            if not has_any_data:
                return redirect('onboarding')
    except UserProfile.DoesNotExist:
        # Ensure profile exists, then redirect
        UserProfile.objects.get_or_create(user=self.request.user if hasattr(self, 'request') else request.user)
        return redirect('onboarding')

    # Process recurring transactions
    process_user_recurring_transactions(request.user)

    # Base QuerySet
    expenses = Expense.objects.filter(user=request.user).order_by('-date')
    
    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings from lists
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range takes precedence
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
        
        # Reset lists for UI clarity since we are in range mode
        selected_years = []
        selected_months = []
        
        trend_title = _("Expenses Trend (Custom Range)")
    else:
        # Default to current month/year ONLY on initial land (no params)
        if not request.GET and not (selected_years or selected_months):
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)
            
        if len(selected_months) == 1 and len(selected_years) == 1:
            trend_title = _("Daily Expenses for %(month)s/%(year)s") % {'month': selected_months[0], 'year': selected_years[0]}
        else:
            trend_title = _("Monthly Expenses Trend")

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
        
    # Income Logic (Mirroring Expense Filters)
    incomes = Income.objects.filter(user=request.user)
    if start_date or end_date:
        if start_date:
            incomes = incomes.filter(date__gte=start_date)
        if end_date:
            incomes = incomes.filter(date__lte=end_date)
    else:
        if selected_years:
            incomes = incomes.filter(date__year__in=selected_years)
        if selected_months:
            incomes = incomes.filter(date__month__in=selected_months)
    
    total_income = incomes.aggregate(Sum('base_amount'))['base_amount__sum'] or 0
    all_dates = Expense.objects.filter(user=request.user).dates('date', 'year', order='DESC')
    years = sorted(list(set([d.year for d in all_dates] + [datetime.now().year])), reverse=True)
    all_categories = Expense.objects.filter(user=request.user).values_list('category', flat=True).distinct().order_by('category')

    # 1. Category Chart Data (Distribution) & Summary Table
    # We need to fetch raw values and merge them in Python to handle whitespace duplicates
    raw_category_data = expenses.values('category').annotate(total=Sum('base_amount'))
    
    # Process and merge duplicates
    merged_category_map = {}
    for item in raw_category_data:
        # Strip whitespace to normalize
        cat_name = item['category'].strip()
        amount = float(item['total'])
        
        if cat_name in merged_category_map:
            merged_category_map[cat_name] += amount
        else:
            merged_category_map[cat_name] = amount
            
    # Convert back to list of dicts for template/charts, sorted by total
    # This replaces the DB-ordered queryset with a sorted list
    category_data = [
        {'category': cat, 'total': amt} 
        for cat, amt in merged_category_map.items()
    ]
    category_data.sort(key=lambda x: x['total'], reverse=True)

    # Compute limits and usage per category for chart display
    # Compute limits and usage per category for chart display
    category_limits = []
    # Optimization: Pre-fetch all categories for the user to avoid N+1 queries in the loop
    user_categories = {c.name: c for c in Category.objects.filter(user=request.user)}

    for item in category_data:
        cat_name = item['category']
        cat_obj = user_categories.get(cat_name)
        
        limit = float(cat_obj.limit) if (cat_obj and cat_obj.limit) else None
        
        used_percent = round((item['total'] / limit * 100), 1) if limit else None
        category_limits.append({
            'name': cat_name,
            'total': item['total'],
            'limit': limit,
            'used_percent': used_percent,
        })
    
    categories = [item['category'] for item in category_data]
    category_amounts = [item['total'] for item in category_data]
    
    # 2. Time Trend (Stacked) Data
    
    # Determine Labels (X-Axis)
    # Determine Labels (X-Axis)
    if start_date or end_date:
        # For custom range, if range < 60 days, show daily. Else monthly.
        # Simple heuristic: Always show daily for custom range for now, or let logic decide.
        # Let's stick to: if explicit month selected -> daily. If range -> daily (usually granular).
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    elif len(selected_months) == 1 and len(selected_years) == 1:
        # Daily view
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    else:
        # Monthly view
        trend_qs = expenses.annotate(period=TruncMonth('date'))
        date_format = '%b %Y'

    # Aggregate by Period AND Category for Stacking
    stacked_data = trend_qs.values('period', 'category').annotate(total=Sum('base_amount')).order_by('period')
    
    # Process into Chart.js Datasets
    # 1. Get unique sorted periods
    periods = sorted(list(set(item['period'] for item in stacked_data)))
    trend_labels = [p.strftime(date_format) for p in periods]
    
    # 2. Build datasets map: { 'CategoryA': [0, 10, 0...], 'CategoryB': ... }
    # Initialize with zeros for all unique NORMALIZED categories found in expenses
    normalized_all_categories = sorted(list(merged_category_map.keys()))
    dataset_map = { cat: [0] * len(periods) for cat in normalized_all_categories }
    
    for item in stacked_data:
        p_idx = periods.index(item['period'])
        # Strip to match our normalized keys
        cat = item['category'].strip()
        if cat in dataset_map:
            dataset_map[cat][p_idx] += float(item['total']) # Add += in case multiple unstripped cats map to same striped cat in same period
            
    # 3. Convert map to list of dataset objects for Chart.js
    trend_datasets = []
    # Define a color palette (Light Blue, Blue Green, Prussian Blue, Honey Yellow, Orange)
    colors = ['#219EBC', '#023047', '#8ECAE6', '#FFB703', '#0575E6']
    
    for i, (cat, data) in enumerate(dataset_map.items()):
        # Only include non-zero datasets
        if sum(data) > 0:
             trend_datasets.append({
                 'label': cat,
                 'data': data,
                 'backgroundColor': colors[i % len(colors)],
                 'borderRadius': 2
             })

    # 3. Top 5 Expenses
    top_expenses_qs = expenses.order_by('-base_amount')[:5]
    top_labels = [
        (e.description.decode('utf-8', errors='replace') if isinstance(e.description, bytes) else str(e.description))[:20] + '...' 
        if len(str(e.description)) > 20 else str(e.description) 
        for e in top_expenses_qs
    ]
    top_amounts = [float(e.base_amount) for e in top_expenses_qs]

    # --- NEW: Income vs Expenses Trend Data ---
    # Re-use the truncation logic determined above
    if start_date or end_date or (len(selected_months) == 1 and len(selected_years) == 1):
        trunc_func = TruncDay
    else:
        trunc_func = TruncMonth
        
    inc_trend = incomes.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('base_amount')).order_by('period')
    exp_trend = expenses.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('base_amount')).order_by('period')
    
    # Merge periods
    inc_periods = set(i['period'] for i in inc_trend)
    exp_periods = set(e['period'] for e in exp_trend)
    all_periods_sorted = sorted(list(inc_periods.union(exp_periods)))
    
    ie_labels = [p.strftime(date_format) for p in all_periods_sorted]
    
    # Optimization: Use dict lookup instead of filter inside loop
    inc_map = {i['period']: float(i['total']) for i in inc_trend}
    exp_map = {e['period']: float(e['total']) for e in exp_trend}
    
    ie_income_data = [inc_map.get(p, 0.0) for p in all_periods_sorted]
    ie_expense_data = [exp_map.get(p, 0.0) for p in all_periods_sorted]
    ie_savings_data = [inc_map.get(p, 0.0) - exp_map.get(p, 0.0) for p in all_periods_sorted]

    # --- NEW: Payment Method Distribution ---
    raw_payment_data = expenses.values('payment_method').annotate(total=Sum('base_amount')).order_by('payment_method')
    payment_map = {}
    for item in raw_payment_data:
        pm_name = item['payment_method'] or 'Unknown'
        payment_map[pm_name] = float(item['total'])
    
    # Sort by total desc
    sorted_payment_items = sorted(payment_map.items(), key=lambda x: x[1], reverse=True)
    payment_labels = [item[0] for item in sorted_payment_items]
    payment_data = [item[1] for item in sorted_payment_items]


    # 4. Summary Stats
    total_expenses = expenses.aggregate(Sum('base_amount'))['base_amount__sum'] or 0
    transaction_count = expenses.count()
    top_category = category_data[0] if category_data else None
    
    savings = total_income - total_expenses

    # --- NEW: Savings Projection (Linear Extrapolation) ---
    current_date = date.today()
    current_year = current_date.year
    current_month = current_date.month 

    # 1. Calculate YTD Savings (Strictly for current year, regardless of filters)
    ytd_income = Income.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
    ytd_expenses = Expense.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
    ytd_savings = ytd_income - ytd_expenses
    
    projected_savings = 0
    
    # Only project if we have data and positive savings
    if ytd_savings > 0:
        # Avoid division by zero if it's January (month 1)
        # Actually, even in Jan, months_passed is 1. So we are good.
        months_passed = current_month
        avg_monthly_savings = ytd_savings / months_passed
        
        months_remaining = 12 - months_passed
        projected_additional = avg_monthly_savings * months_remaining
        
        projected_savings = ytd_savings + projected_additional
    else:
        # If savings are negative or zero, projection is effectively "0" or "current state"
        # We might handle this in template
        projected_savings = 0

    # Calculate MoM Changes ONLY if exactly one year and one month are selected
    prev_month_data = None
    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            sel_year = int(selected_years[0])
            sel_month = int(selected_months[0])
            
            # Calculate previous month and year
            if sel_month == 1:
                prev_month = 12
                prev_year = sel_year - 1
            else:
                prev_month = sel_month - 1
                prev_year = sel_year

            prev_expenses = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            prev_income = Income.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            prev_savings = prev_income - prev_expenses

            def calc_pct(current, previous):
                if previous == 0:
                    return None
                return ((current - previous) / previous) * 100

            prev_month_data = {
                'income_pct': calc_pct(total_income, prev_income),
                'expense_pct': calc_pct(total_expenses, prev_expenses),
                'savings_pct': calc_pct(savings, prev_savings),
            }
            # Add absolute values for template display
            for key in list(prev_month_data.keys()):
                val = prev_month_data[key]
                if val is not None:
                    prev_month_data[f'{key}_abs'] = abs(val)
        except (ValueError, IndexError):
            pass

    # Prepare display labels for the template
    display_year = None
    display_month = None
    
    if len(selected_years) == 1:
        display_year = selected_years[0]
        
    if len(selected_months) == 1:
        try:
            m_idx = int(selected_months[0])
            display_month = _(calendar.month_name[m_idx])
        except (ValueError, IndexError):
            pass

    # NEW: Calculate Previous/Next Month URLs
    prev_month_url = None
    next_month_url = None

    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            curr_year = int(selected_years[0])
            curr_month = int(selected_months[0])
            
            # Previous Month
            if curr_month == 1:
                pm = 12
                py = curr_year - 1
            else:
                pm = curr_month - 1
                py = curr_year
            
            # Next Month
            if curr_month == 12:
                nm = 1
                ny = curr_year + 1
            else:
                nm = curr_month + 1
                ny = curr_year

            # Construct Query String (Preserve Categories)
            base_qs = []
            for c in selected_categories:
                base_qs.append(f'category={c}')
            
            qs_prev = base_qs + [f'year={py}', f'month={pm}']
            qs_next = base_qs + [f'year={ny}', f'month={nm}']
            
            prev_month_url = f"{reverse('home')}?{'&'.join(qs_prev)}"
            next_month_url = f"{reverse('home')}?{'&'.join(qs_next)}"
            
        except ValueError:
            pass
    
    # --- Emotional Feedback / Insights Logic (Enhanced) ---
    
    insights = []
    
    # helper for streaks
    def get_monthly_savings_status(u, y, m):
        inc = Income.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        exp = Expense.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        return inc > exp

    # Construct date params for deep linking
    date_params = ""
    for y in selected_years:
        date_params += f"&year={y}"
    for m in selected_months:
        date_params += f"&month={m}"

    # helper for category links
    def link_cats(cats):
        links_html = format_html_join(
            mark_safe(', '),
            '<a href="{}" class="alert-link text-decoration-underline">{}</a>',
            ((reverse('expense-list') + f"?category={c}{date_params}", c) for c in cats[:2])
        )
        if len(cats) > 2:
            return format_html('{}, etc.', links_html)
        return links_html

    # 0. Anomaly Detection (Spending Spike)
    # Only if viewing current month (or default view)
    is_current_month_view = False
    now = datetime.now()
    if not request.GET or (len(selected_months) == 1 and str(now.month) in selected_months and str(now.year) in selected_years):
         is_current_month_view = True
    
    if is_current_month_view and total_expenses > 0:
        # Calculate last 3 months average
        last_3_months_total = 0
        months_counted = 0
        for i in range(1, 4):
            # Calculate past month/year
            y = now.year
            m = now.month - i
            while m < 1:
                m += 12
                y -= 1
            
            m_total = Expense.objects.filter(user=request.user, date__year=y, date__month=m).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            if m_total > 0:
                last_3_months_total += m_total
                months_counted += 1
        
        if months_counted > 0:
            avg_past_spend = last_3_months_total / months_counted
            
            # Project current month
            days_in_month = calendar.monthrange(now.year, now.month)[1]
            days_passed = now.day
            if days_passed > 0:
                projected_spend = (float(total_expenses) / days_passed) * days_in_month
                avg_past_spend_float = float(avg_past_spend)
                
                if projected_spend > avg_past_spend_float * 1.25 and float(total_expenses) > 1000: # 25% Higher + Min Threshold
                    pct_higher = int(((projected_spend - avg_past_spend_float) / avg_past_spend_float) * 100)
                    insights.append({
                        'type': 'warning',
                        'icon': 'graph-up-arrow',
                        'title': _('Traffic Alert 🚦'),
                        'message': _("You're pacing %(pct_higher)s%% higher than usual. Slow down to stay on track!") % {'pct_higher': pct_higher},
                        'allow_share': False
                    })

    # 1. Budget Warnings (High Priority)

    over_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and c['used_percent'] > 100]
    near_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and 90 <= c['used_percent'] <= 100]
    
    # Check savings rate for "Softener" context
    savings_rate = (savings / total_income * 100) if total_income > 0 else 0
    
    if over_budget_cats:
        cats_str = link_cats(over_budget_cats)
        
        if savings_rate >= 20:
            # Contextualized Warning for High Savers
            msg = format_html(_("Even strong months have leaks. You crossed limits in {cats_str} — catching this keeps you on track."), cats_str=cats_str)
        else:
            # Standard Coaching Warning - "Warning" type (Yellow) instead of Danger (Red) for empathy
            msg = format_html(_("⚠️ Budget crossed in {cats_str} — let’s rebalance to stay safe."), cats_str=cats_str)

        insights.append({
            'type': 'warning', # Changed from danger
            'icon': 'exclamation-octagon-fill',
            'title': _('Budget Breached'),
            'message': msg,
            'allow_share': False
        })
    elif near_budget_cats:
        cats_str = link_cats(near_budget_cats)
        insights.append({
            'type': 'warning',
            'icon': 'exclamation-triangle-fill',
            'title': _('Approaching Limit'),
            'message': format_html(_("Heads up! You're close to overspending on {cats_str}."), cats_str=cats_str),
            'allow_share': False
        })

    # 2. Wins & Cause-Based Praise (Specific & Celebratory)
    if prev_month_data:
        # Calculate Category Savings (Cause of the win)
        # We need prev month category breakdown
        prev_cat_qs = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).values('category').annotate(total=Sum('base_amount'))
        prev_cat_map = {item['category'].strip(): float(item['total']) for item in prev_cat_qs}
        
        savings_contributors = []
        for cat, curr_total in merged_category_map.items():
            prev_total = prev_cat_map.get(cat, 0)
            if prev_total > curr_total:
                diff = prev_total - curr_total
                if diff > 100: # Threshold to mention
                    savings_contributors.append((cat, diff))
        savings_contributors.sort(key=lambda x: x[1], reverse=True)
        top_savers = [c[0] for c in savings_contributors[:2]]
        
        # Savings Win
        if total_income > 0 and savings > 0:
            savings_rate = (savings / total_income) * 100
            if savings_rate >= 20:
                msg_text = _("You've saved %(savings_rate)s%% of your income this month.") % {'savings_rate': f"{savings_rate:.0f}"}
                share_text = _("I saved %(savings_rate)s%% of my income this month using TrackMyRupee! 🏆") % {'savings_rate': f"{savings_rate:.0f}"}
                
                if top_savers:
                    cats_link = link_cats(top_savers)
                    msg = format_html(_("{msg_text} You spent less on {cats_link} — that's where the magic happened."), msg_text=msg_text, cats_link=cats_link)
                else:
                    msg = msg_text

                insights.append({
                    'type': 'success',
                    'icon': 'trophy-fill',
                    'title': _('Super Saver Status! 🏆'),
                    'message': msg,
                    'allow_share': True,
                    'share_text': share_text
                })
            elif prev_month_data['savings_pct'] and prev_month_data['savings_pct'] > 0:
                 insights.append({
                    'type': 'success',
                    'icon': 'graph-up-arrow',
                    'title': _('Momentum Building 🚀'),
                    'message': _("Your savings grew by %(savings_pct_abs)s%% vs last month. You're getting better at this!") % {'savings_pct_abs': f"{prev_month_data['savings_pct_abs']:.0f}"},
                    'allow_share': True,
                    'share_text': _("My savings grew by %(savings_pct_abs)s%% this month! 🚀 via TrackMyRupee") % {'savings_pct_abs': f"{prev_month_data['savings_pct_abs']:.0f}"}
                })
        
        # Expense Control Win (if we haven't already praised savings)
        if len(insights) == 0: 
            if prev_month_data['expense_pct'] and prev_month_data['expense_pct'] < -5:
                 msg_text = _("You've cut spending by %(expense_pct_abs)s%%.") % {'expense_pct_abs': f"{prev_month_data['expense_pct_abs']:.0f}"}
                 share_text = _("I cut my spending by %(expense_pct_abs)s%% this month! 👍 via TrackMyRupee") % {'expense_pct_abs': f"{prev_month_data['expense_pct_abs']:.0f}"}
                 
                 if top_savers:
                     cats_link = link_cats(top_savers)
                     msg = format_html(_("{msg_text} {cats_link} saw the biggest drops."), msg_text=msg_text, cats_link=cats_link)
                 else:
                     msg = msg_text
                 
                 insights.append({
                    'type': 'success',
                    'icon': 'check-circle-fill',
                    'title': _('You’re in Control 👍'),
                    'message': msg,
                    'allow_share': True,
                    'share_text': _("I cut my spending by %(expense_pct_abs)s%% this month! 👍 via TrackMyRupee") % {'expense_pct_abs': f"{prev_month_data['expense_pct_abs']:.0f}"}
                })

    # 3. Streak & Identity (Reassuring / Habit Forming)
    # Only calculate if current status is good
    if savings > 0 and len(selected_years) == 1 and len(selected_months) == 1:
        streak = 1 # Current month counts
        check_to_go = 5 # check max 5 months back
        curr_y_calc, curr_m_calc = int(selected_years[0]), int(selected_months[0])
        
        for i in range(check_to_go):
            # Go back one month
            if curr_m_calc == 1:
                curr_m_calc = 12
                curr_y_calc -= 1
            else:
                curr_m_calc -= 1
            
            if get_monthly_savings_status(request.user, curr_y_calc, curr_m_calc):
                streak += 1
            else:
                break
        
        if streak > 1:
            insights.append({
                'type': 'info', # Use Info for "Identity/Streak"
                'icon': 'fire',
                'title': _('On a Roll!'),
                'message': _("🔥 This is your %(streak)s month in a row staying under budget.") % {'streak': streak},
                'allow_share': True,
                'share_text': _("🔥 I've stayed under budget for %(streak)s months in a row! via TrackMyRupee") % {'streak': streak}
            })

    # 4. Fallback
    if not insights and savings > 0:
        insights.append({
            'type': 'info',
            'icon': 'piggy-bank-fill',
            'title': _('In the Green'),
            'message': _("You've saved %(savings)s so far. Keep it up!") % {'savings': savings},
            'allow_share': False
        })
    elif not insights:
        insights.append({
            'type': 'secondary',
            'icon': 'stars',
            'title': _('Fresh Start'),
            'message': _("Small steps today lead to big results tomorrow. Let's track some expenses!"),
            'allow_share': False
        })

    # Limit to top 2 insights to avoid clutter
    insights = insights[:2]

    # Check for onboarding (True if user has NO data at all)
    has_any_data = Expense.objects.filter(user=request.user).exists() or Income.objects.filter(user=request.user).exists()

    # Logic for "Year in Review" Banner
    show_year_in_review = False
    year_in_review_year = None
    if has_any_data:
        # Show last year's review from Jan to Oct
        # Show this year's review in Nov/Dec
        if now.month >= 11:
            year_in_review_year = now.year
        else:
            year_in_review_year = now.year - 1
            
        if year_in_review_year:
            show_year_in_review = Expense.objects.filter(user=request.user, date__year=year_in_review_year).exists()

    context = {
        'is_new_user': not has_any_data,
        'insights': insights[::-1],
        'total_income': total_income,
        'savings': savings,
        'recent_transactions': expenses.order_by('-date')[:5],
        'categories': categories,
        'category_amounts': category_amounts,
        'category_data': category_data, # Passing full queryset for the summary table
        'category_limits': category_limits,
        'trend_labels': trend_labels,
        'trend_datasets': trend_datasets,
        'trend_title': trend_title,
        'top_labels': top_labels,
        'top_amounts': top_amounts,
        # New Context
        'ie_labels': ie_labels,
        'ie_income_data': ie_income_data,
        'ie_expense_data': ie_expense_data,
        'ie_savings_data': ie_savings_data,
        'payment_labels': payment_labels,
        'payment_data': payment_data,
        'years': years,
        'all_categories': all_categories,
        'selected_years': selected_years,
        'selected_months': selected_months,
        'selected_year': display_year,    # NEW: For template display labels
        'selected_month': display_month,  # NEW: For template display labels
        'selected_categories': selected_categories,
        'months_list': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'total_expenses': total_expenses,
        'transaction_count': transaction_count,
        'top_category': top_category,
        'projected_savings': projected_savings, # NEW
        'start_date': start_date,
        'end_date': end_date,
        'prev_month_data': prev_month_data,
        'prev_month_url': prev_month_url,
        'next_month_url': next_month_url,
        'show_tutorial': not request.user.profile.has_seen_tutorial or request.GET.get('tour') == 'true',
        'has_any_budget': any((c.get('limit') or 0) > 0 for c in category_limits),
        'show_year_in_review': show_year_in_review,
        'year_in_review_year': year_in_review_year,
    }
    return render(request, 'home.html', context)

@login_required
def complete_tutorial(request):
    if request.method == 'POST':
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        profile.has_seen_tutorial = True
        profile.save()
        return JsonResponse({'success': True})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def upload_view(request):
    """
    Upload view with year selection enforcement.
    """
    
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        selected_year = int(request.POST.get('year'))
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                
                if not rows:
                    continue

                # Search for the header row index
                header_row_index = -1
                header_cols = []
                
                for i, row in enumerate(rows[:10]):
                    if not row: continue
                    row_values = [str(val).strip().title() if val is not None else "" for val in row]
                    if 'Date' in row_values and 'Amount' in row_values and 'Description' in row_values:
                        header_row_index = i
                        header_cols = row_values
                        break
                
                if header_row_index == -1:
                    print(f"Skipping sheet {sheet_name}: Could not find header row.")
                    continue

                # Map column indices
                col_map = {col: idx for idx, col in enumerate(header_cols) if col}
                required_columns = ['Date', 'Amount', 'Description', 'Category']
                
                if not all(col in col_map for col in required_columns):
                    print(f"Skipping sheet {sheet_name}: Missing required columns.")
                    continue

                # Process data rows
                for row_data in rows[header_row_index + 1:]:
                    if not any(row_data): continue # Skip empty rows
                    
                    # Parse date
                    date_val = row_data[col_map['Date']]
                    if date_val is None:
                        continue
                        
                    date_obj = None
                    if isinstance(date_val, str):
                        formats = ['%d %b %Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %B %Y', '%d %b', '%d-%b', '%d %B']
                        for fmt in formats:
                            try:
                                parsed_date = datetime.strptime(date_val.strip(), fmt).date()
                                date_obj = parsed_date.replace(year=selected_year)
                                break
                            except ValueError:
                                continue
                        if not date_obj:
                            continue
                    elif isinstance(date_val, (datetime, date)):
                        date_obj = date_val if isinstance(date_val, date) else date_val.date()
                        try:
                            date_obj = date_obj.replace(year=selected_year)
                        except ValueError:
                            date_obj = date_obj.replace(day=28, year=selected_year)
                    else:
                        continue # Unsupported date type

                    # Get other fields
                    amount = row_data[col_map['Amount']]
                    description = row_data[col_map['Description']]
                    category = row_data[col_map['Category']] if 'Category' in col_map else None
                    
                    if amount is None or description is None:
                        continue

                    category_obj = None
                    if category:
                        category_name = str(category).strip()
                        if category_name:
                            category_obj, _ = Category.objects.get_or_create(user=request.user, name=category_name)

                    Expense.objects.get_or_create(
                        user=request.user,
                        date=date_obj,
                        amount=float(amount) if not isinstance(amount, float) else amount,
                        description=str(description),
                        category=category_obj.name if category_obj else "Others"
                    )
            return redirect('home')
        except Exception as e:
            print(f"Error processing file: {e}")
            traceback.print_exc()
            pass

    # Context for year dropdown
    current_year = datetime.now().year
    years = range(current_year, current_year - 5, -1)
    
    return render(request, 'upload.html', {'years': years, 'current_year': current_year})

class ExpenseListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Expense
    template_name = 'expenses/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20

    def get_queryset(self):
        queryset = Expense.objects.filter(user=self.request.user).order_by('-date')
        
        # Filtering
        selected_years = self.request.GET.getlist('year')
        selected_months = self.request.GET.getlist('month')
        selected_categories = self.request.GET.getlist('category')
        search_query = self.request.GET.get('search')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        # Remove empty strings from lists
        selected_years = [y for y in selected_years if y]
        selected_months = [m for m in selected_months if m]
        selected_categories = [c for c in selected_categories if c]
        
        # Date Range Logic (Precedence over Year/Month)
        if start_date or end_date:
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__lte=end_date)
        else:
            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query  # Don't check categories as we might want defaults even if cat is selected? No, usually filters are additive.
            )
            
            # If no year/month/search filters, default to current month/year
            # (ignoring category here might be debated, but typically if I just filter 'Food', I might want all time or current month? 
            #  The dashboard logic defaults to current month if no year/month. Let's stick to that.)
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            if selected_years:
                queryset = queryset.filter(date__year__in=selected_years)
            
            if selected_months:
                queryset = queryset.filter(date__month__in=selected_months)

        if selected_categories:
            queryset = queryset.filter(category__in=selected_categories)
        
        # Filter by Payment Method
        payment_method = self.request.GET.get('payment_method')
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        if search_query:
            queryset = queryset.filter(description__icontains=search_query)
            
        # Sorting
        sort_by = self.request.GET.get('sort')
        if sort_by == 'amount_asc':
            queryset = queryset.order_by('amount')
        elif sort_by == 'amount_desc':
            queryset = queryset.order_by('-amount')
        # Default is already '-date' from line 961, so valid fallback.
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('base_amount'))['base_amount__sum'] or 0

        # Get unique years and categories for validation
        user_expenses = Expense.objects.filter(user=self.request.user)
        years_dates = user_expenses.dates('date', 'year', order='DESC')
        years = sorted(list(set([d.year for d in years_dates] + [datetime.now().year])), reverse=True)
        # Python-side deduplication to handle whitespace variants (e.g. "Goa" vs "Goa ")
        raw_used_categories = user_expenses.values_list('category', flat=True)
        raw_defined_categories = Category.objects.filter(user=self.request.user).values_list('name', flat=True)
        all_cats = set([c.strip() for c in raw_used_categories if c and c.strip()]) | set([c.strip() for c in raw_defined_categories if c and c.strip()])
        categories = sorted(list(all_cats), key=str.lower)
        
        context['years'] = years
        context['categories'] = categories
        context['months_list'] = [(i, calendar.month_name[i]) for i in range(1, 13)]
        
        # Determine selected year for UI
        # Determine selected year for UI
        year_param = self.request.GET.get('year')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        context['start_date'] = start_date
        context['end_date'] = end_date
        
        if start_date or end_date:
            context['selected_years'] = []
            context['selected_months'] = []
            context['selected_categories'] = []
        else:
            selected_years = self.request.GET.getlist('year')
            selected_months = self.request.GET.getlist('month')
            selected_categories = self.request.GET.getlist('category')
            search_query = self.request.GET.get('search')
            
            # Remove empty strings
            selected_years = [y for y in selected_years if y]
            selected_months = [m for m in selected_months if m]
            selected_categories = [c for c in selected_categories if c]

            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query 
                # (ignoring category here as well to match get_queryset)
            )

            # Mirror default logic from get_queryset
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            context['selected_years'] = selected_years
            context['selected_months'] = selected_months
            context['selected_categories'] = selected_categories
            
        return context

class ExpenseCreateView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'expenses/expense_form.html'

    def get(self, request, *args, **kwargs):
        # We need to wrap the formset to pass 'user' to the form constructor
        ExpenseFormSet = modelformset_factory(Expense, form=ExpenseForm, extra=1, can_delete=True)
        # Pass user to form kwargs using formset_factory's form_kwargs (requires Django 4.0+)
        # For older Django or modelformset, we might need a custom formset or curry the form.
        # Simpler approach: Use a lambda or partial, but modelformset_factory creates a class.
        
        # Actually, best way for modelformset with custom init args is to override BaseFormSet or manually iterate.
        # But simpler hack: Set the widget choices in the view by iterating forms? No, new forms need it.
        
        # Let's use form_kwargs in the formset initialization if supported.
        # Django 1.9+ supports form_kwargs in formset constructor.
        
        initial_data = [{'date': datetime.now().date(), 'currency': request.user.profile.currency} for _ in range(1)]
        formset = ExpenseFormSet(queryset=Expense.objects.none(), initial=initial_data, form_kwargs={'user': request.user})
        next_url = request.GET.get('next', '')
        
        # Get top 5 frequent categories for this user
        frequent_categories = Expense.objects.filter(user=request.user).values('category').annotate(count=Count('category')).order_by('-count')[:5]
        frequent_category_names = [item['category'] for item in frequent_categories]

        return render(request, self.template_name, {
            'formset': formset, 
            'next_url': next_url,
            'frequent_categories': frequent_category_names
        })

    def post(self, request, *args, **kwargs):
        ExpenseFormSet = modelformset_factory(Expense, form=ExpenseForm, extra=1, can_delete=True)
        formset = ExpenseFormSet(request.POST, form_kwargs={'user': request.user})
        if formset.is_valid():
            try:
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.user = request.user
                    instance.save()
                
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('expense-list')
            except IntegrityError:
                messages.error(request, _("Duplicate record found! You already have this expense recorded for this date."))
                return render(request, self.template_name, {'formset': formset})
        return render(request, self.template_name, {'formset': formset})

class ExpenseUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'expenses/expense_form.html'
    success_url = reverse_lazy('expense-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        
        # Get top 5 frequent categories for this user
        frequent_categories = Expense.objects.filter(user=self.request.user).values('category').annotate(count=Count('category')).order_by('-count')[:5]
        context['frequent_categories'] = [item['category'] for item in frequent_categories]
        
        return context

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This expense entry already exists.")
            return self.form_invalid(form)

    def get_queryset(self):
        # Ensure user can only edit their own expenses
        return Expense.objects.filter(user=self.request.user)

class ExpenseBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        expense_ids = request.POST.getlist('expense_ids')
        if not expense_ids:
            messages.error(request, 'No expenses selected for deletion.')
            return redirect('expense-list')
            
        # Filter by IDs and ensuring they belong to the current user for security
        expenses_to_delete = Expense.objects.filter(id__in=expense_ids, user=request.user)
        deleted_count = expenses_to_delete.count()
        
        if deleted_count > 0:
            expenses_to_delete.delete()
            messages.success(request, f'{deleted_count} expenses deleted successfully.')
        else:
            messages.warning(request, 'No valid expenses found to delete.')
            
        return redirect('expense-list')

class ExpenseBulkUpdateView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        expense_ids = request.POST.getlist('expense_ids')
        category = request.POST.get('bulk_category')
        payment_method = request.POST.get('bulk_payment_method')
        
        if not expense_ids:
            messages.error(request, _('No expenses selected for update.'))
            return redirect('expense-list')
            
        update_data = {}
        if category:
            update_data['category'] = category
        if payment_method:
            update_data['payment_method'] = payment_method
            
        if not update_data:
            messages.warning(request, _('No fields selected to update.'))
            return redirect('expense-list')
            
        # Filter by IDs and ensure they belong to the current user
        expenses_to_update = Expense.objects.filter(id__in=expense_ids, user=request.user)
        updated_count = expenses_to_update.count()
        
        if updated_count > 0:
            expenses_to_update.update(**update_data)
            messages.success(request, _(f'{updated_count} expenses updated successfully.'))
        else:
            messages.warning(request, _('No valid expenses found to update.'))
            
        return redirect('expense-list')

class ExpenseDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Expense
    template_name = 'expenses/expense_confirm_delete.html'
    success_url = reverse_lazy('expense-list')

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)
    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)

class CategoryListView(LoginRequiredMixin, generic.ListView):
    model = Category
    template_name = 'expenses/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

    def get_queryset(self):
        queryset = Category.objects.filter(user=self.request.user).order_by('name')
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        
        # Nudge context for upgrade banner (use is_plus/is_pro to respect subscription expiry)
        profile = self.request.user.profile
        if not profile.is_pro:
            total_categories = Category.objects.filter(user=self.request.user).count()
            if profile.is_plus:
                limit = 10
                upgrade_tier = 'PRO'
            else:
                limit = 5
                upgrade_tier = 'PLUS'
            context['reached_limit'] = total_categories >= limit
            context['current_count'] = total_categories
            context['limit'] = limit
            context['nudge_current'] = total_categories
            context['nudge_limit'] = limit
            context['nudge_feature_name'] = 'categories'
            context['nudge_upgrade_tier'] = upgrade_tier
            context['nudge_at_limit'] = total_categories >= limit
        
        context['bootstrap_icons'] = BOOTSTRAP_ICONS
        return context

class CategoryCreateView(LoginRequiredMixin, generic.CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')
 
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bootstrap_icons'] = BOOTSTRAP_ICONS
        # Check Limits
        current_count = Category.objects.filter(user=self.request.user).count()
        limit = 5 # Free
        if self.request.user.profile.is_plus:
            limit = 10
        if self.request.user.profile.is_pro:
            limit = float('inf')

        context['reached_limit'] = current_count >= limit
        context['category_limit'] = limit
        return context

    def form_valid(self, form):
        # Double check limit on POST
        current_count = Category.objects.filter(user=self.request.user).count()
        limit = 5
        if self.request.user.profile.is_plus:
            limit = 10
        if self.request.user.profile.is_pro:
            limit = float('inf')
        
        if current_count >= limit:
            messages.error(self.request, _("Category limit reached. Please upgrade to add more."))
            return redirect('category-create')

        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, _("This category already exists."))
            return self.form_invalid(form)

class CategoryUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')
 
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bootstrap_icons'] = BOOTSTRAP_ICONS
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        profile = request.user.profile
        if not profile.is_pro:
            limit = 10 if profile.is_plus else 5
            # We use the same ordering as in ExpenseForm for consistency
            categories = list(Category.objects.filter(user=request.user).order_by('id'))
            try:
                index = categories.index(obj)
                if index >= limit:
                    messages.error(request, _("This category is locked due to your current plan limits. Upgrade to edit."))
                    return redirect('category-list')
            except ValueError:
                pass
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)
    
    def form_valid(self, form):
        try:
            # Store old name to update related expenses
            old_name = self.get_object().name
            response = super().form_valid(form)
            new_name = self.object.name
            
            if old_name != new_name:
                Expense.objects.filter(user=self.request.user, category=old_name).update(category=new_name)
                
            return response
        except IntegrityError:
            messages.error(self.request, "This category already exists.")
            return self.form_invalid(form)

class CategoryDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Category
    template_name = 'expenses/category_confirm_delete.html'
    success_url = reverse_lazy('category-list')

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)

@login_required
def export_expenses(request):
    """
    Export expenses to CSV based on current filters.
    """
    # Check Limits
    if not request.user.profile.is_plus:
        messages.error(request, "Export is available on Plus and Pro plans.")
        return redirect('pricing')

    expenses = Expense.objects.filter(user=request.user).order_by('-date')

    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    search_query = request.GET.get('search')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range Logic (Precedence over Year/Month)
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
    else:
        # Check if any specific filter is active
        has_active_filters = (
            selected_years or 
            selected_months or 
            search_query
        )
        
        # If no year/month/search filters, default to current month/year
        if not has_active_filters:
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
    if search_query:
        expenses = expenses.filter(description__icontains=search_query)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

    writer = csv.writer(response)
    writer.writerow([_('Date'), _('Category'), _('Description'), _('Currency'), _('Amount'), _('Amount (Base)')])

    for expense in expenses:
        writer.writerow([
            expense.date, 
            expense.category, 
            expense.description, 
            expense.currency, 
            expense.amount, 
            expense.base_amount
        ])

    return response

# --------------------
# Income Views
# --------------------

class IncomeListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Income
    template_name = 'expenses/income_list.html'
    context_object_name = 'incomes'
    paginate_by = 20

    def get_queryset(self):
        queryset = Income.objects.filter(user=self.request.user).order_by('-date')
        
        # Default dates (Current Year)
        today = timezone.localdate()
        default_start = today.replace(month=1, day=1)
        default_end = today.replace(month=12, day=31)

        # Date Filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        source = self.request.GET.get('source')

        # Check if we have ANY filter params. If not, apply default dates.
        if not date_from and not date_to and not source:
             self.date_from = default_start.isoformat()
             self.date_to = default_end.isoformat()
             queryset = queryset.filter(date__gte=default_start, date__lte=default_end)
        else:
            # We have some filters (or user explicitly cleared them? - tricky part about "reset")
            # If user wants to "clear" filters, they usually submit empty strings.
            # But the requirement says "default start date...". Usually implies initial load.
            if date_from:
                queryset = queryset.filter(date__gte=date_from)
                self.date_from = date_from
            else:
                self.date_from = ''
            
            if date_to:
                queryset = queryset.filter(date__lte=date_to)
                self.date_to = date_to
            else:
                self.date_to = ''

        # Source Filter
        if source:
            queryset = queryset.filter(source__icontains=source)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import CURRENCY_CHOICES
        context['currency_choices'] = CURRENCY_CHOICES
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        
        context['filter_form'] = {
            'date_from': getattr(self, 'date_from', ''),
            'date_to': getattr(self, 'date_to', ''),
            'source': self.request.GET.get('source', ''),
        }
        return context

class IncomeCreateView(LoginRequiredMixin, generic.CreateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, _("This income entry already exists."))
            return self.form_invalid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context


class IncomeUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, _("This income entry already exists."))
            return self.form_invalid(form)

class IncomeDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Income
    template_name = 'expenses/income_confirm_delete.html'
    success_url = reverse_lazy('income-list')

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)



class CalendarView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = datetime.now()
        
        # Get year/month from URL or default to current
        year = self.kwargs.get('year', today.year)
        month = self.kwargs.get('month', today.month)
        
        # Validate year/month
        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            year = today.year
            month = today.month

        # Calculate prev/next month for navigation
        if month == 1:
            prev_month_date = date(year - 1, 12, 1)
        else:
            prev_month_date = date(year, month - 1, 1)
            
        if month == 12:
            next_month_date = date(year + 1, 1, 1)
        else:
            next_month_date = date(year, month + 1, 1)

        # Get search query
        search_query = self.request.GET.get('search', '')

        # Base filters
        expense_filters = Q(user=self.request.user, date__year=year, date__month=month)
        income_filters = Q(user=self.request.user, date__year=year, date__month=month)
        
        if search_query:
            # Filter expenses by description or category
            expense_filters &= (Q(description__icontains=search_query) | Q(category__icontains=search_query))
            # Filter income by source or description
            income_filters &= (Q(source__icontains=search_query) | Q(description__icontains=search_query))

        # Get Expense and Income Data for the month
        expenses = Expense.objects.filter(expense_filters).values('date').annotate(total=Sum('base_amount'))
        
        incomes = Income.objects.filter(income_filters).values('date').annotate(total=Sum('base_amount'))
        
        # Map data for easy lookup by day
        # Keys are integers (day of month)
        expense_map = {e['date'].day: e['total'] for e in expenses}
        income_map = {i['date'].day: i['total'] for i in incomes}
        
        # Build Calendar Grid
        cal = calendar.Calendar(firstweekday=6) # Start on Sunday
        month_days = cal.monthdayscalendar(year, month)
        
        # Transform into a list of weeks, where each day is an object
        calendar_data = []
        for week in month_days:
            week_data = []
            for day in week:
                if day == 0:
                    week_data.append(None) # Empty slot
                else:
                    week_data.append({
                        'day': day,
                        'income': income_map.get(day, 0),
                        'expense': expense_map.get(day, 0),
                    })
            calendar_data.append(week_data)
        
        
        # Calculate totals for the month to show net savings
        total_monthly_expense = sum(item['total'] for item in expenses) or 0
        total_monthly_income = sum(item['total'] for item in incomes) or 0
        month_net_savings = total_monthly_income - total_monthly_expense

        context['calendar_data'] = calendar_data
        context['current_year'] = year
        context['current_month'] = month
        context['month_name'] = date_format(date(year, month, 1), 'F')
        context['month_net_savings'] = month_net_savings
        context['prev_year'] = prev_month_date.year
        context['prev_month'] = prev_month_date.month
        context['next_year'] = next_month_date.year
        context['next_month'] = next_month_date.month
        context['search_query'] = search_query
        
        return context


class BudgetDashboardView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/budget_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = date.today()
        
        month_param = self.request.GET.get('month')
        year_param = self.request.GET.get('year')
        
        month = int(month_param) if month_param else today.month
        year = int(year_param) if year_param else today.year
        
        # Ensure context variables for filters are correct
        context['current_month'] = month
        context['current_year'] = year
        
        categories = Category.objects.filter(user=user)
        budget_data = []
        
        total_budget = 0
        categorized_spent = 0
        
        # Calculate total spending across ALL expenses for the month
        grand_total_spent = Expense.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        ).aggregate(Total=Sum('base_amount'))['Total'] or 0

        for category in categories:
            spent = Expense.objects.filter(
                user=user,
                category=category.name,
                date__year=year,
                date__month=month
            ).aggregate(Total=Sum('base_amount'))['Total'] or 0
            
            percentage = (spent / category.limit * 100) if category.limit and category.limit > 0 else 0
            
            budget_data.append({
                'category': category,
                'spent': spent,
                'limit': category.limit,
                'percentage': min(percentage, 100),
                'actual_percentage': percentage,
                'remaining': (category.limit - spent) if category.limit and spent <= category.limit else 0,
                'over_budget': (spent - category.limit) if category.limit and spent > category.limit else 0
            })
            
            if category.limit:
                total_budget += category.limit
            categorized_spent += spent
            
        context.update({
            'budget_data': budget_data,
            'total_budget': total_budget,
            'total_spent': grand_total_spent,
            'total_remaining': (total_budget - grand_total_spent) if total_budget > grand_total_spent else 0,
            'over_budget_amount': (grand_total_spent - total_budget) if grand_total_spent > total_budget else 0,
            'total_percentage': min((grand_total_spent / total_budget * 100), 100) if total_budget else 0,
            'actual_total_percentage': (grand_total_spent / total_budget * 100) if total_budget else 0,
            'month_name': date(year, month, 1).strftime('%B'),
        })

        # MoM Calculation for Budget Dashboard
        if month == 1:
            prev_month = 12
            prev_year = year - 1
        else:
            prev_month = month - 1
            prev_year = year

        prev_spent = Expense.objects.filter(
            user=user,
            date__year=prev_year,
            date__month=prev_month
        ).aggregate(Total=Sum('base_amount'))['Total'] or 0

        if prev_spent > 0:
            context['spent_mom_pct'] = ((grand_total_spent - prev_spent) / prev_spent) * 100
            context['spent_mom_pct_abs'] = abs(context['spent_mom_pct'])
        else:
            context['spent_mom_pct'] = None
            context['spent_mom_pct_abs'] = None

        context.update({
            'current_month': month,
            'current_year': year,
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'years': range(today.year - 2, today.year + 2),
        })
        return context

# --------------------
# Recurring Transaction Views
# --------------------

class RecurringTransactionListView(LoginRequiredMixin, ListView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_list.html'
    context_object_name = 'recurring_transactions'
    filter_expenses_only = True

    def get_queryset(self):
        queryset = RecurringTransaction.objects.filter(user=self.request.user)
        if self.filter_expenses_only:
            queryset = queryset.filter(transaction_type='EXPENSE')
        queryset = queryset.order_by('-created_at')
        
        # Filter by Category
        categories = self.request.GET.getlist('category')
        if categories:
            queryset = queryset.filter(category__in=categories)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_transactions = self.object_list
        today = date.today()
        
        # Categories for filter
        user_transactions = RecurringTransaction.objects.filter(user=self.request.user)
        categories = user_transactions.values_list('category', flat=True).distinct().order_by('category')
        # Filter out None/Empty if any
        categories = [c for c in categories if c]
        
        context['categories'] = categories
        context['selected_categories'] = self.request.GET.getlist('category')
        
        # Split into Active and Cancelled
        # We sort active subs by creation date to determine which ones are locked
        active_subs = [t for t in all_transactions if t.is_active]
        active_subs.sort(key=lambda x: x.created_at or x.id) # Fallback to ID if created_at is null
        
        profile = self.request.user.profile
        limit = float('inf')
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 0
            
        for i, sub in enumerate(active_subs):
            sub.is_locked = i >= limit
            
        cancelled_subs = [t for t in all_transactions if not t.is_active]
        
        # Calculate Totals (Monthly & Yearly)
        total_monthly = 0
        total_yearly = 0
        
        for sub in active_subs:
            amount = sub.base_amount
            if sub.frequency == 'DAILY':
                total_monthly += amount * 30
                total_yearly += amount * 365
            elif sub.frequency == 'WEEKLY':
                total_monthly += amount * 4
                total_yearly += amount * 52
            elif sub.frequency == 'MONTHLY':
                total_monthly += amount
                total_yearly += amount * 12
            elif sub.frequency == 'YEARLY':
                total_monthly += amount / 12
                total_yearly += amount

        # Identify "Renewing Soon" (This Month)
        renewing_soon = []
        renewals_count = 0
        
        # Helper to find next date relative to today
        for sub in active_subs:
            # Calculate next occurrence
            next_date = sub.start_date
            
            # For simpler logic, we reset the year/month to current to check basic interval
            # But for accurate "days until", we need better logic:
            
            if sub.frequency == 'DAILY':
                next_date = today + timedelta(days=1)
            elif sub.frequency == 'WEEKLY':
                # Find days ahead
                days_ahead = (sub.start_date.weekday() - today.weekday()) % 7
                if days_ahead == 0 and today > sub.start_date: # if today is the day, but older start
                     days_ahead = 7
                elif days_ahead == 0 and today == sub.start_date: # exact match today
                     days_ahead = 0
                else: 
                     # If start_date was future, we wait. If past, we find next.
                     # Simplified: just next occurrence of that weekday
                     if days_ahead <= 0: days_ahead += 7
                
                # Correction: Standard logic to find next matching weekday
                days_ahead = (sub.start_date.weekday() - today.weekday()) 
                if days_ahead <= 0: # Target day already happened this week or is today
                    days_ahead += 7
                next_date = today + timedelta(days=days_ahead)
                
            elif sub.frequency == 'MONTHLY':
                # Occurs on sub.start_date.day every month
                # If today.day > start_date.day, it's next month.
                # If today.day <= start_date.day, it's this month.
                try:
                    if today.day > sub.start_date.day:
                        # Next month
                        month = today.month + 1
                        year = today.year
                        if month > 12:
                            month = 1
                            year += 1
                        next_date = date(year, month, sub.start_date.day)
                    else:
                        # This month
                        next_date = date(today.year, today.month, sub.start_date.day)
                except ValueError: 
                    # Handle end of month issues (e.g. 31st) - simplified to 1st of next-next month
                    next_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1)

            elif sub.frequency == 'YEARLY':
                try:
                    this_year_date = date(today.year, sub.start_date.month, sub.start_date.day)
                    if today > this_year_date:
                        next_date = date(today.year + 1, sub.start_date.month, sub.start_date.day)
                    else:
                        next_date = this_year_date
                except ValueError:
                    next_date = date(today.year, 2, 28)

            # Annotate object
            sub.annotated_next_date = next_date
            sub.annotated_days_until = (next_date - today).days
            
            # Determine urgency
            is_renewing = False
            if sub.transaction_type == 'EXPENSE':
                if sub.annotated_days_until <= 30: # Show mostly anything coming up soon
                     is_renewing = True
            
            if is_renewing:
                renewing_soon.append(sub)
                renewals_count += 1
            
            # Sort renewing soon by days until
            renewing_soon.sort(key=lambda x: x.annotated_days_until)

        context.update({
            'active_subs': active_subs,
            'cancelled_subs': cancelled_subs,
            'renewing_soon': renewing_soon,
            'renewals_count': renewals_count,
            'total_monthly_cost': total_monthly,
            'total_yearly_cost': total_yearly,
        })
        
        # Nudge context for upgrade banner (use is_plus/is_pro to respect subscription expiry)
        profile = self.request.user.profile
        active_count = RecurringTransaction.objects.filter(user=self.request.user, is_active=True).count()
        limit = float('inf')
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 0
            
            if profile.is_plus:
                upgrade_tier = 'PRO'
            else:
                upgrade_tier = 'PLUS'
            context['nudge_current'] = active_count
            context['nudge_limit'] = limit
            context['nudge_feature_name'] = 'recurring transactions'
            context['nudge_upgrade_tier'] = upgrade_tier
            context['nudge_at_limit'] = active_count >= limit
            # Free users: always show nudge (they have 0 limit)
            # Plus users: show when >= 60% of 3 = 2+
            if limit == 0:
                context['show_nudge'] = True
            else:
                context['show_nudge'] = active_count >= max(1, int(limit * 0.6))
        
        context['is_limit_reached'] = active_count >= limit
        context['current_limit'] = limit
        
        return context

class RecurringTransactionManageView(RecurringTransactionListView):
    template_name = 'expenses/recurring_transaction_manage.html'
    filter_expenses_only = False

class RecurringTransactionCreateView(LoginRequiredMixin, CreateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def dispatch(self, request, *args, **kwargs):
        # Prevent access if limit reached
        profile = request.user.profile
        active_count = RecurringTransaction.objects.filter(user=request.user, is_active=True).count()
        limit = float('inf')
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 0
            
        if active_count >= limit:
            messages.error(request, _("Subscription limit reached for your plan. Please upgrade to add more."))
            return redirect('recurring-list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check Limits
        current_count = RecurringTransaction.objects.filter(user=self.request.user, is_active=True).count()
        limit = 0 # Free
        if self.request.user.profile.is_plus:
            limit = 3
        if self.request.user.profile.is_pro:
            limit = float('inf')

        if current_count >= limit:
             messages.error(self.request, _("Recurring Transaction limit reached (%(limit)s). Please upgrade.") % {'limit': limit})
             return redirect('pricing')
             
        form.instance.user = self.request.user
        messages.success(self.request, _('Recurring transaction created successfully.'))
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

class RecurringTransactionUpdateView(LoginRequiredMixin, UpdateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        profile = request.user.profile
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 0
            # Same ordering as in RecurringTransactionListView
            active_subs = list(RecurringTransaction.objects.filter(user=request.user, is_active=True).order_by('created_at', 'id'))
            try:
                index = active_subs.index(obj)
                if index >= limit:
                    messages.error(request, _("This subscription is locked due to your current plan limits. Upgrade to edit."))
                    return redirect('recurring-list')
            except ValueError:
                # If sub is not active, we still check among all recurring transactions?
                # Usually "locked" applies to active ones that stopped being processed.
                pass
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check if is_active changed from True to False
        if 'is_active' in form.changed_data and not form.cleaned_data['is_active']:
            # Cancellation detected
            obj = self.get_object() # get current object
            # Calculate yearly saving
            amount = obj.amount
            if obj.frequency == 'DAILY':
                yearly_saving = amount * 365
            elif obj.frequency == 'WEEKLY':
                yearly_saving = amount * 52
            elif obj.frequency == 'MONTHLY':
                yearly_saving = amount * 12
            else: # YEARLY
                yearly_saving = amount
            
            # Assuming currency symbol is available in request or we use generic. 
            # We can use the profile currency if available, or just a generic prompt. 
            # User request used '₹', but code uses {{ currency_symbol }} in template.
            # We'll try to fetch user currency or default.
            currency = '₹'
            if hasattr(self.request.user, 'userprofile'):
                currency = self.request.user.userprofile.currency
                
            messages.success(self.request, _("You just saved %(currency)s%(amount)s/year 🎉") % {'currency': currency, 'amount': f"{yearly_saving:,.0f}"})
        else:
            messages.success(self.request, _('Recurring transaction updated successfully.'))
            
        return super().form_valid(form)

class RecurringTransactionDeleteView(LoginRequiredMixin, DeleteView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_confirm_delete.html' # Added template_name for consistency
    success_url = reverse_lazy('recurring-list')

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def form_valid(self, form):
        # Calculate savings
        obj = self.object
        amount = obj.amount
        if obj.frequency == 'DAILY':
            yearly_saving = amount * 365
        elif obj.frequency == 'WEEKLY':
            yearly_saving = amount * 52
        elif obj.frequency == 'MONTHLY':
            yearly_saving = amount * 12
        else: # YEARLY
            yearly_saving = amount
            
        currency = '₹'
        if hasattr(self.request.user, 'userprofile'):
            currency = self.request.user.userprofile.currency
            
        messages.success(self.request, _("You just saved %(currency)s%(amount)s/year 🎉") % {'currency': currency, 'amount': f"{yearly_saving:,.0f}"})
        return super().form_valid(form)

class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model = User
    success_url = reverse_lazy('landing')
    template_name = 'expenses/account_confirm_delete.html'

    def get_object(self, queryset=None):
        return self.request.user

    def form_valid(self, form):
        user = self.get_object()
        logout(self.request) # Log out before deleting
        user.delete()
        messages.success(self.request, "Your account has been deleted successfully.")
        return redirect(self.success_url)

class CurrencyUpdateView(LoginRequiredMixin, UpdateView):
    model = UserProfile
    fields = ['currency']
    template_name = 'expenses/currency_settings.html'
    success_url = reverse_lazy('currency-settings')

    def get_object(self, queryset=None):
        profile, created = UserProfile.objects.get_or_create(user=self.request.user)
        return profile

    def form_valid(self, form):
        old_currency = self.get_object().currency
        new_currency = form.cleaned_data.get('currency')
        from django.db import IntegrityError
        
        response = super().form_valid(form)
        
        if old_currency != new_currency:
            # Re-normalize all transactions
            user = self.request.user
            skipped_count = 0
            for model in [Expense, Income, RecurringTransaction]:
                transactions = model.objects.filter(user=user)
                for tx in transactions:
                    try:
                        tx.save() # This will trigger the new save() logic with the new base_currency
                    except IntegrityError:
                        skipped_count += 1
                        continue
            
            if skipped_count > 0:
                messages.warning(self.request, f'Currency preference updated. {skipped_count} transactions were skipped due to potential duplication.')
            else:
                messages.success(self.request, 'Currency preference updated successfully.')
        else:
            messages.success(self.request, 'Currency preference updated successfully.')
            
        return response

class LanguageUpdateView(LoginRequiredMixin, UpdateView):
    model = UserProfile
    form_class = LanguageUpdateForm
    template_name = 'expenses/language_settings.html'
    success_url = reverse_lazy('language-settings')

    def get_object(self, queryset=None):
        profile, created = UserProfile.objects.get_or_create(user=self.request.user)
        return profile

    def form_valid(self, form):
        from django.conf import settings
        from django.utils import translation
        lang = form.cleaned_data.get('language')
        translation.activate(lang)
        messages.success(self.request, 'Language preference updated successfully.')
        
        response = super().form_valid(form)
        response.set_cookie(settings.LANGUAGE_COOKIE_NAME, lang)
        return response

class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = ProfileUpdateForm
    template_name = 'expenses/profile_settings.html'
    success_url = reverse_lazy('profile-settings')

    def get_object(self):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Profile Settings'
        context['is_social_user'] = SocialAccount.objects.filter(user=self.request.user).exists()
        
        # Handle Year in Review visibility
        now = timezone.now()
        has_any_data = Expense.objects.filter(user=self.request.user).exists() or Income.objects.filter(user=self.request.user).exists()
        show_year_in_review = False
        year_in_review_year = None
        
        if has_any_data:
            # Logic: Show previous until Oct, then current
            if now.month >= 11:
                year_in_review_year = now.year
            else:
                year_in_review_year = now.year - 1
                
            if year_in_review_year:
                show_year_in_review = Expense.objects.filter(user=self.request.user, date__year=year_in_review_year).exists()
                
        context['show_year_in_review'] = show_year_in_review
        context['year_in_review_year'] = year_in_review_year
        
        return context

    def form_valid(self, form):
        messages.success(self.request, "Profile updated successfully.")
        return super().form_valid(form)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    """
    # Clear any existing messages (e.g. from previous logout)
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        # Manually set the backend to allow login without authentication
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
        return redirect('home')
    except User.DoesNotExist:
        messages.error(request, "Demo user not setup. Please contact admin.")
        return redirect('account_login')

class PricingView(TemplateView):
    template_name = 'expenses/pricing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['RAZORPAY_KEY_ID'] = settings.RAZORPAY_KEY_ID
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans_monthly'] = {p.tier: p for p in plans.filter(duration='MONTHLY')}
        context['plans_yearly'] = {p.tier: p for p in plans.filter(duration='YEARLY')}
        # Backward compat: default "plans" uses yearly
        context['plans'] = context['plans_yearly']
        return context

def ping(request):
    return HttpResponse("Pong", status=200)

class ContactView(View):
    template_name = 'contact.html'
    
    # Spam protection settings
    RATE_LIMIT_HOURLY = 3
    RATE_LIMIT_DAILY = 10
    MIN_MESSAGE_LENGTH = 10
    
    # Common spam patterns
    SPAM_KEYWORDS = [
        'precio', 'price check', 'buy now', 'click here', 'earn money',
        'viagra', 'casino', 'lottery', 'prize', 'congratulations',
        'limited offer', 'act now', 'online pharmacy', 'weight loss',
        'make money fast', 'work from home', 'investment opportunity',
        'hola, quería saber', 'please kindly', 'dear friend'
    ]
    
    # Disposable email domains
    DISPOSABLE_DOMAINS = [
        'tempmail.com', 'guerrillamail.com', '10minutemail.com',
        'throwaway.email', 'maildrop.cc', 'mailinator.com',
        'trashmail.com', 'yopmail.com', 'getnada.com'
    ]

    def get(self, request):
        form = ContactForm()
        return render(request, self.template_name, {'form': form})
    
    def _get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def _check_rate_limit(self, ip):
        """Check if IP has exceeded rate limits"""
        
        hourly_key = f'contact_hourly_{ip}'
        daily_key = f'contact_daily_{ip}'
        
        hourly_count = cache.get(hourly_key, 0)
        daily_count = cache.get(daily_key, 0)
        
        if hourly_count >= self.RATE_LIMIT_HOURLY:
            return False, "Too many submissions. Please try again in an hour."
        
        if daily_count >= self.RATE_LIMIT_DAILY:
            return False, "Daily submission limit reached. Please try again tomorrow."
        
        # Increment counters
        cache.set(hourly_key, hourly_count + 1, 3600)  # 1 hour
        cache.set(daily_key, daily_count + 1, 86400)   # 24 hours
        
        return True, None
    
    def _is_spam_content(self, text):
        """Check if text contains spam patterns"""
        text_lower = text.lower()
        
        # Check for URLs (most spam contains links)
        if 'http://' in text_lower or 'https://' in text_lower or 'www.' in text_lower:
            return True, "Messages with URLs are not allowed."
        
        # Check for spam keywords
        for keyword in self.SPAM_KEYWORDS:
            if keyword in text_lower:
                return True, "Your message was flagged as potential spam."
        
        # Check for excessive caps (> 50% uppercase)
        if len(text) > 20:
            caps_count = sum(1 for c in text if c.isupper())
            if caps_count / len(text) > 0.5:
                return True, "Please don't use excessive capitalization."
        
        # Check message length
        if len(text.strip()) < self.MIN_MESSAGE_LENGTH:
            return True, "Please provide a more detailed message."
        
        return False, None
    
    def _is_disposable_email(self, email):
        """Check if email is from a disposable domain"""
        domain = email.split('@')[-1].lower()
        return domain in self.DISPOSABLE_DOMAINS

    def post(self, request):
        form = ContactForm(request.POST)
        
        # This handles validations for all fields including reCAPTCHA (if configured)
        if not form.is_valid():
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'form': form})

        # Get cleaned data
        data = form.cleaned_data
        name = data.get('name')
        email = data.get('email')
        subject = data.get('subject')
        message = data.get('message')
        honeypot = data.get('website')
        
        # Layer 1: Honeypot check
        if honeypot:
            # Silently reject spam bots - don't reveal honeypot was triggered
            messages.success(request, "Your message has been sent! We'll get back to you shortly.")
            return redirect('contact')
        
        # Layer 2: Rate limiting
        client_ip = self._get_client_ip(request)
        rate_ok, rate_msg = self._check_rate_limit(client_ip)
        if not rate_ok:
            messages.error(request, rate_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 3: Content filtering
        is_spam, spam_msg = self._is_spam_content(subject + ' ' + message)
        if is_spam:
            messages.error(request, spam_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 4: Email validation
        if self._is_disposable_email(email):
            messages.error(request, "Please use a permanent email address.")
            return render(request, self.template_name, {'form': form})
        
        # Layer 5: reCAPTCHA verification is handled by form.is_valid()
        
        # All checks passed - send email
        full_message = f"""
        New Contact Form Submission:
        
        Name: {name}
        Email: {email}
        Subject: {subject}
        IP: {client_ip}
        
        Message:
        {message}
        """

        try:
            send_mail(
                subject=f"Contact Form: {subject}",
                message=full_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=['track.my.rupee.app@gmail.com'],
                fail_silently=False,
            )
            messages.success(request, "Your message has been sent! We'll get back to you shortly.")
            return redirect('contact')
        except Exception:
            # Log error if possible
            messages.error(request, "Something went wrong. Please try again later.")
            return render(request, self.template_name, {'form': form})

@login_required
def predict_category_view(request):
    """
    AJAX view to predict category based on description.
    Only available on PRO tier (AI Insights).
    """
    if request.method == 'GET':
        # Gating: Only for Pro users
        if not request.user.profile.is_pro:
            return JsonResponse({'category': None, 'error': 'AI Insights is a Pro feature.'}, status=403)

        description = request.GET.get('description', '').strip()
        if not description:
             return JsonResponse({'category': None})
        
        category = predict_category_ai(description, request.user)
        return JsonResponse({'category': category})

# --------------------
# Notification Views
# --------------------

class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'expenses/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(user=self.request.user, is_read=False).count()
        return context

@login_required
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, "All notifications marked as read.")
        return redirect('notification-list')
    return redirect('notification-list')

@login_required
def mark_single_notification_read(request, pk):
    try:
        notification = Notification.objects.get(pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)

@csrf_exempt
def trigger_notifications(request):
    """
    HTTP endpoint to trigger notifications via external cron service (e.g. cron-job.org).
    Secured by a secret key in the URL params: ?secret=YOUR_CRON_SECRET
    """
    secret = request.GET.get('secret')
    
    # Check against dedicated CRON_SECRET
    if not secret or secret != settings.CRON_SECRET:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    try:
        call_command('send_notifications')
        return JsonResponse({'success': True, 'message': 'Notifications triggered successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    if request.method == 'POST':
        notification = get_object_or_404(Notification, pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@csrf_exempt
def trigger_lifecycle_emails(request):
    """
    HTTP endpoint to trigger lifecycle drip emails via external cron service.
    Secured by a secret key in the URL params: ?secret=YOUR_CRON_SECRET
    """
    secret = request.GET.get('secret')

    if not secret or secret != settings.CRON_SECRET:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        call_command('send_lifecycle_emails')
        return JsonResponse({'success': True, 'message': 'Lifecycle emails triggered successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


class AnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Check for Pro Access
        if not hasattr(user, 'profile') or not user.profile.is_pro:
            context['is_locked'] = True
            return context
        
        today = timezone.now().date()
        
        # Get Available Years from User Data
        expense_years = list(Expense.objects.filter(user=user).dates('date', 'year').values_list('date__year', flat=True))
        income_years = list(Income.objects.filter(user=user).dates('date', 'year').values_list('date__year', flat=True))
        available_years = sorted(list(set(expense_years + income_years)), reverse=True)
        
        if not available_years:
            available_years = [today.year]
            
        # Determine Selected Year
        selected_year_str = self.request.GET.get('year')
        if selected_year_str and selected_year_str.isdigit():
            selected_year = int(selected_year_str)
        else:
            selected_year = today.year
            
        context['available_years'] = available_years
        context['selected_year'] = selected_year
        context['is_current_year'] = (selected_year == today.year)
        
        # 1. Monthly Trends (Selected Year)
        labels = []
        income_data = []
        expense_data = []
        balance_rate_data = []
        
        # Determine the start and end date for the selected year
        start_date = date(selected_year, 1, 1)
        end_date = date(selected_year, 12, 31)
        
        # Fetch data grouped by Month for the selected year
        monthly_income = Income.objects.filter(
            user=user, date__gte=start_date, date__lte=end_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('base_amount')).order_by('month')
        
        monthly_expenses = Expense.objects.filter(
            user=user, date__gte=start_date, date__lte=end_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('base_amount')).order_by('month')
        
        # Merge data into a map {date: {income: 0, expense: 0}}
        data_map = {}
        
        # Initialize map with all 12 months to ensure 0s for missing months
        # Iterate from start_date to today month by month
        curr = start_date
        while curr <= today:
            d = curr.replace(day=1)
            data_map[d] = {'income': 0, 'expense': 0}
            # Move to next month
            # Carefully handle month increment
            next_month = curr.month + 1
            next_year = curr.year
            if next_month > 12:
                next_month = 1
                next_year += 1
            curr = date(next_year, next_month, 1)

        # Fill with DB data
        # Fill with DB data
        for item in monthly_income:
            if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['income'] = float(item['total'])
                
        for item in monthly_expenses:
             if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['expense'] = float(item['total'])
                
        # Sort and prepare lists
        sorted_keys = sorted(data_map.keys())
        # Limit to last 12 months if while loop went over
        sorted_keys = sorted_keys[-12:]
        
        for k in sorted_keys:
            labels.append(date_format(k, 'M Y'))
            inc = data_map[k]['income']
            exp = data_map[k]['expense']
            income_data.append(inc)
            expense_data.append(exp)
            
            # Balance Rate = (Income - Expense) / Income * 100
            if inc > 0:
                rate = ((inc - exp) / inc) * 100
            else:
                rate = 0
            balance_rate_data.append(round(rate, 1))

        context['chart_labels'] = labels
        context['income_data'] = income_data
        context['expense_data'] = expense_data
        context['balance_rate_data'] = balance_rate_data
        
        # 2. Category Breakdown (Selected Year)
        category_stats = Expense.objects.filter(
            user=user, date__year=selected_year
        ).values('category').annotate(total=Sum('base_amount')).order_by('-total')
        
        cat_labels = [_(x['category']) for x in category_stats]
        cat_data = [float(x['total']) for x in category_stats]
        
        context['cat_labels'] = cat_labels
        context['cat_data'] = cat_data
        
        # 3. Key Metrics (YTD / Full Year depending on selection)
        if selected_year == today.year:
            # For current year, limit to today so future recurring entries don't skew YTD
            ytd_income_agg = Income.objects.filter(user=user, date__year=selected_year, date__lte=today).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            ytd_expense_agg = Expense.objects.filter(user=user, date__year=selected_year, date__lte=today).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        else:
            # For past years, show the full year's total
            ytd_income_agg = Income.objects.filter(user=user, date__year=selected_year).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            ytd_expense_agg = Expense.objects.filter(user=user, date__year=selected_year).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        
        context['total_income_ytd'] = ytd_income_agg
        context['total_expense_ytd'] = ytd_expense_agg
        context['total_balance_ytd'] = ytd_income_agg - ytd_expense_agg
        
        if ytd_income_agg > 0:
            context['avg_balance_rate'] = round(((ytd_income_agg - ytd_expense_agg) / ytd_income_agg) * 100, 1)
        else:
            context['avg_balance_rate'] = 0
            
        # ---------------------------------------------------------
        # 4. Sankey Data (Income -> Expenses/Savings) YTD
        # ---------------------------------------------------------
        sankey_data = []
        if ytd_income_agg > 0 or ytd_expense_agg > 0:
            top_cats = list(category_stats[:8])
            other_cats = list(category_stats[8:])
            
            for cat in top_cats:
                sankey_data.append([_('Income'), _(cat['category']), float(cat['total'])])
                
            if other_cats:
                other_total = sum(float(cat['total']) for cat in other_cats)
                if other_total > 0:
                    sankey_data.append([_('Income'), _('Other Expenses'), other_total])
                    
            savings = ytd_income_agg - ytd_expense_agg
            if savings > 0:
                sankey_data.append([_('Income'), _('Savings'), float(savings)])
            elif savings < 0:
                # If deficit, flow deficit into Income so Google Chart balances it as money source
                sankey_data.append([_('Deficit (Overspending)'), _('Income'), abs(float(savings))])
                
        context['sankey_data'] = sankey_data
            
        # ---------------------------------------------------------
        # 5. Cashflow Forecasting (Next 6 Months)
        # ---------------------------------------------------------
        
        # A. Calculate Historical Average (Last 3 completed months)
        # We need completed months to avoid partial data skewing
        avg_income = 0
        avg_expense = 0
        months_counted = 0
        
        # Iterate back 3 months
        for i in range(1, 4):
            # Calculate past month/year
            # e.g. If now is May, we want April, March, Feb
            past_date = today.replace(day=1) - timedelta(days=20*i) # rough jump back
            # normalize to 1st of that month
            past_date = past_date.replace(day=1)
            
            # Re-calculate cleanly
            # If today is 2023-05-15
            # i=1 -> 2023-04-01
            # i=2 -> 2023-03-01
            # i=3 -> 2023-02-01
            year = today.year
            month = today.month - i
            while month < 1:
                month += 12
                year -= 1
            
            p_income = Income.objects.filter(user=user, date__year=year, date__month=month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            p_expense = Expense.objects.filter(user=user, date__year=year, date__month=month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            
            avg_income += p_income
            avg_expense += p_expense
            months_counted += 1
            
        if months_counted > 0:
            avg_income /= months_counted
            avg_expense /= months_counted

        # ---------------------------------------------------------
        # 5. Financial Health Score & Insights
        # ---------------------------------------------------------
        balance_rate = context.get('avg_balance_rate', 0)
        health_score = 0
        health_label = ""
        health_color = ""
        insights = []
        
        # Currency symbol for formatting
        currency = user.profile.currency if hasattr(user, 'profile') else '₹'
        savings_ytd = float(ytd_income_agg) - float(ytd_expense_agg)
        income_f = float(ytd_income_agg)
        expense_f = float(ytd_expense_agg)
        
        # Top category context (used across insights)
        top_cat_name = cat_labels[0] if cat_data else None
        top_cat_pct = 0
        if top_cat_name and income_f > 0:
            top_cat_pct = (cat_data[0] / income_f) * 100
        
        # Score Logic with dynamic messages
        if balance_rate < 0:
            health_score = 10
            health_label = _("At Risk")
            health_color = "danger"
            overspend_pct = abs(balance_rate)
            msg = _("You spent %(currency)s%(expense)s against %(currency)s%(income)s income (%(pct)s%% overspend).") % {
                'currency': currency,
                'expense': f"{expense_f:,.0f}",
                'income': f"{income_f:,.0f}",
                'pct': f"{overspend_pct:.0f}",
            }
            if top_cat_name:
                msg += " " + _("Review '%(cat)s' which accounts for %(cat_pct)s%% of your spending.") % {
                    'cat': escape(top_cat_name),
                    'cat_pct': f"{top_cat_pct:.0f}",
                }
            insights.append(msg)
        elif balance_rate < 10:
            health_score = 40
            health_label = _("Caution")
            health_color = "warning"
            msg = _("You saved %(currency)s%(savings)s out of %(currency)s%(income)s (%(rate)s%%).") % {
                'currency': currency,
                'savings': f"{savings_ytd:,.0f}",
                'income': f"{income_f:,.0f}",
                'rate': f"{balance_rate:.0f}",
            }
            if top_cat_name and cat_data:
                potential_saving = cat_data[0] * 0.2
                msg += " " + _("Cutting '%(cat)s' by 20%% could save %(currency)s%(amount)s more.") % {
                    'cat': escape(top_cat_name),
                    'currency': currency,
                    'amount': f"{potential_saving:,.0f}",
                }
            insights.append(msg)
        elif balance_rate < 30:
            health_score = 70
            health_label = _("Stable")
            health_color = "info"
            gap_to_30 = (0.30 * income_f) - savings_ytd if income_f > 0 else 0
            msg = _("You saved %(currency)s%(savings)s out of %(currency)s%(income)s (%(rate)s%%).") % {
                'currency': currency,
                'savings': f"{savings_ytd:,.0f}",
                'income': f"{income_f:,.0f}",
                'rate': f"{balance_rate:.0f}",
            }
            if gap_to_30 > 0:
                msg += " " + _("Just %(currency)s%(gap)s more to hit the 30%% savings benchmark.") % {
                    'currency': currency,
                    'gap': f"{gap_to_30:,.0f}",
                }
            insights.append(msg)
        else:
            health_score = 95
            health_label = _("Wealth Builder")
            health_color = "success"
            insights.append(
                _("You saved %(currency)s%(savings)s out of %(currency)s%(income)s (%(rate)s%%). That's top-tier financial discipline.") % {
                    'currency': currency,
                    'savings': f"{savings_ytd:,.0f}",
                    'income': f"{income_f:,.0f}",
                    'rate': f"{balance_rate:.0f}",
                }
            )

        # Category Insight (high concentration warning)
        if cat_data and top_cat_name:
            if income_f > 0:
                try:
                    if top_cat_pct > 30:
                        safe_top_cat = escape(top_cat_name)
                        insights.append(
                            _("'%(cat)s' is consuming %(pct)s%% of your income (%(currency)s%(amount)s). Consider setting a budget limit.") % {
                                'cat': safe_top_cat,
                                'pct': f"{top_cat_pct:.0f}",
                                'currency': currency,
                                'amount': f"{cat_data[0]:,.0f}",
                            }
                        )
                except (ValueError, TypeError):
                    pass

        # Spending Trend Projection (uses 3-month avg calculated earlier)
        if avg_expense > 0 and expense_f > 0 and selected_year == today.year:
            avg_expense_f = float(avg_expense)
            # Compare current year monthly average to historical
            months_elapsed = today.month
            current_monthly_avg = expense_f / months_elapsed if months_elapsed > 0 else 0
            if current_monthly_avg > avg_expense_f * 1.15:  # 15% higher than historical
                pct_increase = ((current_monthly_avg - avg_expense_f) / avg_expense_f) * 100
                next_month_name = (today.replace(day=28) + timedelta(days=4)).strftime('%B')
                projected = current_monthly_avg
                insights.append(
                    _("Spending is trending %(pct)s%% above your 3-month average. %(month)s may reach %(currency)s%(projected)s at this pace.") % {
                        'pct': f"{pct_increase:.0f}",
                        'month': next_month_name,
                        'currency': currency,
                        'projected': f"{projected:,.0f}",
                    }
                )
        # ---------------------------------------------------------
        # 5b. Health Breakdown Panel (Expandable Metrics)
        # ---------------------------------------------------------
        health_breakdown = []
        
        # 1. Savings Rate
        savings_rate_val = round(balance_rate, 1)
        if savings_rate_val >= 20:
            sr_status, sr_icon = 'success', 'check-circle-fill'
        elif savings_rate_val > 0:
            sr_status, sr_icon = 'warning', 'exclamation-triangle-fill'
        else:
            sr_status, sr_icon = 'danger', 'x-circle-fill'
        health_breakdown.append({
            'label': _('Savings Rate'),
            'value': f"{savings_rate_val}%",
            'status': sr_status,
            'icon': sr_icon,
        })
        
        # 2. Expense Growth (current monthly avg vs 3-month historical avg)
        avg_expense_f = float(avg_expense) if avg_expense > 0 else 0
        months_elapsed = today.month if selected_year == today.year else 12
        current_monthly_avg = expense_f / months_elapsed if months_elapsed > 0 else 0
        
        if avg_expense_f > 0:
            expense_growth_pct = round(((current_monthly_avg - avg_expense_f) / avg_expense_f) * 100, 1)
        else:
            expense_growth_pct = 0
        
        eg_prefix = '+' if expense_growth_pct > 0 else ''
        if expense_growth_pct <= 0:
            eg_status, eg_icon = 'success', 'check-circle-fill'
        elif expense_growth_pct <= 15:
            eg_status, eg_icon = 'warning', 'exclamation-triangle-fill'
        else:
            eg_status, eg_icon = 'danger', 'x-circle-fill'
        health_breakdown.append({
            'label': _('Expense Growth'),
            'value': f"{eg_prefix}{expense_growth_pct}%",
            'status': eg_status,
            'icon': eg_icon,
        })
        
        # 3. Consistency (months with positive savings out of last 10)
        consistency_count = 0
        consistency_total = 10
        for i in range(1, consistency_total + 1):
            c_year = today.year
            c_month = today.month - i
            while c_month < 1:
                c_month += 12
                c_year -= 1
            c_inc = Income.objects.filter(user=user, date__year=c_year, date__month=c_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            c_exp = Expense.objects.filter(user=user, date__year=c_year, date__month=c_month).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
            if c_inc > c_exp and c_inc > 0:
                consistency_count += 1
        
        if consistency_count >= 7:
            cs_status, cs_icon = 'success', 'check-circle-fill'
        elif consistency_count >= 4:
            cs_status, cs_icon = 'warning', 'exclamation-triangle-fill'
        else:
            cs_status, cs_icon = 'danger', 'x-circle-fill'
        health_breakdown.append({
            'label': _('Consistency'),
            'value': f"{consistency_count}/{consistency_total}",
            'status': cs_status,
            'icon': cs_icon,
        })
        
        # 4. Risk Buffer (months of runway = total savings / avg monthly expense)
        if avg_expense_f > 0 and savings_ytd > 0:
            risk_buffer_months = round(savings_ytd / avg_expense_f, 1)
        else:
            risk_buffer_months = 0
        
        if risk_buffer_months >= 6:
            rb_status, rb_icon = 'success', 'check-circle-fill'
        elif risk_buffer_months >= 3:
            rb_status, rb_icon = 'warning', 'exclamation-triangle-fill'
        else:
            rb_status, rb_icon = 'danger', 'x-circle-fill'
        health_breakdown.append({
            'label': _('Risk Buffer'),
            'value': _("%(months)s months") % {'months': f"{risk_buffer_months:.0f}"},
            'status': rb_status,
            'icon': rb_icon,
        })
        
        # Health Summary sentence (combine best + worst)
        status_order = {'success': 2, 'warning': 1, 'danger': 0}
        best = max(health_breakdown, key=lambda x: status_order[x['status']])
        worst = min(health_breakdown, key=lambda x: status_order[x['status']])
        
        summary_parts = []
        if best['status'] == 'success':
            summary_parts.append(_("Your %(label)s is excellent.") % {'label': best['label'].lower()})
        if worst['status'] != 'success' and worst != best:
            summary_parts.append(_("However, %(label)s needs attention (%(value)s).") % {'label': worst['label'].lower(), 'value': worst['value']})
        health_summary = ' '.join(str(p) for p in summary_parts) if summary_parts else ''

        context['health_score'] = health_score
        context['health_label'] = health_label
        context['health_color'] = health_color
        context['insights'] = insights
        context['health_breakdown'] = health_breakdown
        context['health_summary'] = health_summary
            


        # 4. Day-of-Week Spending (Selected Year)
        dow_stats = Expense.objects.filter(
            user=user, date__year=selected_year
        ).annotate(weekday=ExtractWeekDay('date')).values('weekday').annotate(
            total=Sum('base_amount')
        ).order_by('weekday')
        
        # ExtractWeekDay: 1 (Sun) to 7 (Sat)
        dow_labels = [_('Sun'), _('Mon'), _('Tue'), _('Wed'), _('Thu'), _('Fri'), _('Sat')]
        dow_data = [0] * 7
        for item in dow_stats:
            dow_data[item['weekday'] - 1] = float(item['total'])
        
        context['dow_labels'] = dow_labels
        context['dow_data'] = dow_data

        # 5. Cumulative Burn-down (Current Month)
        # Only relevant for the current month/year
        this_month_labels = []
        this_month_spent = []
        this_month_budget_line = []
        this_month_projection = []
        
        if today.year == selected_year:
            # Current Month Cumulative
            days_in_month = calendar.monthrange(today.year, today.month)[1]
            daily_spending = Expense.objects.filter(
                user=user, date__year=today.year, date__month=today.month
            ).values('date__day').annotate(total=Sum('base_amount')).order_by('date__day')
            
            daily_map = {item['date__day']: float(item['total']) for item in daily_spending}
            
            total_monthly_limit = Category.objects.filter(user=user).aggregate(Sum('limit'))['limit__sum'] or 0
            total_monthly_limit = float(total_monthly_limit)
            
            cumulative = 0
            this_month_projection = [None] * days_in_month
            
            for day in range(1, days_in_month + 1):
                this_month_labels.append(str(day))
                
                # Only add actual spent data up to today
                if day <= today.day:
                    cumulative += daily_map.get(day, 0)
                    this_month_spent.append(round(cumulative, 2))
                else:
                    this_month_spent.append(None)
                
                # Ideal line (pro-rated budget)
                if total_monthly_limit > 0:
                    this_month_budget_line.append(round((total_monthly_limit / days_in_month) * day, 2))
                else:
                    this_month_budget_line.append(0)

            # Calculation for Projection (Burn Rate)
            if today.day > 0 and cumulative > 0:
                burn_rate = cumulative / today.day
                proj_cumulative = cumulative
                # Start projection from today
                this_month_projection[today.day - 1] = round(cumulative, 2)
                for day in range(today.day + 1, days_in_month + 1):
                    proj_cumulative += burn_rate
                    this_month_projection[day - 1] = round(proj_cumulative, 2)
        
        context['burn_down_labels'] = this_month_labels
        context['burn_down_spent'] = this_month_spent
        context['burn_down_budget'] = this_month_budget_line
        context['burn_down_projection'] = this_month_projection



        # 7. Recurring vs One-time (Current Month)
        recurring_sum = Expense.objects.filter(
            user=user, 
            date__year=today.year, 
            date__month=today.month,
            description__icontains='(Recurring)'
        ).aggregate(Sum('base_amount'))['base_amount__sum'] or 0
        
        total_sum = Expense.objects.filter(
            user=user, date__year=today.year, date__month=today.month
        ).aggregate(Sum('base_amount'))['base_amount__sum'] or 1 # Avoid div by zero
        
        rec_val = float(recurring_sum)
        one_time_val = float(total_sum) - rec_val
        
        context['recurring_split_labels'] = [_('Recurring'), _('One-time')]
        context['recurring_split_data'] = [rec_val, max(0, one_time_val)]

        return context

# --- Savings Goal Views ---

class SavingsGoalListView(LoginRequiredMixin, ListView):
    model = SavingsGoal
    template_name = 'expenses/goal_list.html'
    context_object_name = 'goals'

    def get_queryset(self):
        return SavingsGoal.objects.filter(user=self.request.user).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        goals = context['goals']
        
        # Calculate total saved across all active goals (in user's base currency, simple sum for MVP)
        total_saved = sum(g.current_amount for g in goals)
        context['total_saved'] = round(total_saved, 2)
        
        # Free users get 1 goal, Plus gets 3, Pro gets unlimited
        goal_count = goals.count()
        profile = self.request.user.profile
        
        limit = float('inf')
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 1
            
        context['can_create_goal'] = goal_count < limit
        
        # Mark goals beyond limit as locked (by creation order)
        # Convert to list to avoid evaluating queryset multiple times and to add transient attribute
        goals_list = list(goals.order_by('created_at', 'id'))
        for i, goal in enumerate(goals_list):
            goal.is_locked = i >= limit
            
        context['goals'] = goals_list
        
        if not profile.is_pro:
            context['nudge_current'] = goal_count
            context['nudge_limit'] = 3 if profile.is_plus else 1
            context['nudge_feature_name'] = 'savings goals'
            context['nudge_upgrade_tier'] = 'PRO' if profile.is_plus else 'PLUS'
            context['nudge_at_limit'] = goal_count >= context['nudge_limit']
            # show nudge when at 60% of limit
            limit_val = 3 if profile.is_plus else 1
            context['show_nudge'] = goal_count >= max(1, int(limit_val * 0.6))
        else:
            context['show_nudge'] = False
                
        return context

class SavingsGoalCreateView(LoginRequiredMixin, CreateView):
    model = SavingsGoal
    form_class = SavingsGoalForm
    template_name = 'expenses/goal_form.html'
    success_url = reverse_lazy('goal-list')

    def dispatch(self, request, *args, **kwargs):
        # Limit check (use is_plus/is_pro to respect subscription expiry)
        profile = request.user.profile
        if not profile.is_pro:
            goals_count = SavingsGoal.objects.filter(user=request.user).count()
            if profile.is_plus:
                if goals_count >= 3:
                    messages.error(request, _("PLUS plan is limited to 3 active Savings Goals. Upgrade to PRO to unlock unlimited goals!"))
                    return redirect('goal-list')
            else:
                if goals_count >= 1:
                    messages.error(request, _("Free plan is limited to 1 active Savings Goal. Upgrade to PLUS for 3 goals, or PRO for unlimited!"))
                    return redirect('goal-list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, _("Savings goal created successfully!"))
        return super().form_valid(form)

class SavingsGoalUpdateView(LoginRequiredMixin, UpdateView):
    model = SavingsGoal
    form_class = SavingsGoalForm
    template_name = 'expenses/goal_form.html'
    success_url = reverse_lazy('goal-list')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        profile = request.user.profile
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 1
            # Same ordering as in SavingsGoalListView
            goals = list(SavingsGoal.objects.filter(user=request.user).order_by('created_at', 'id'))
            try:
                index = goals.index(obj)
                if index >= limit:
                    messages.error(request, _("This savings goal is locked due to your current plan limits. Upgrade to edit."))
                    return redirect('goal-list')
            except ValueError:
                pass
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return SavingsGoal.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, _("Savings goal updated successfully!"))
        return super().form_valid(form)

class SavingsGoalDeleteView(LoginRequiredMixin, DeleteView):
    model = SavingsGoal
    success_url = reverse_lazy('goal-list')

    def get_queryset(self):
        return SavingsGoal.objects.filter(user=self.request.user)
        
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, _("Savings goal deleted successfully."))
        return super().delete(request, *args, **kwargs)

class SavingsGoalDetailView(LoginRequiredMixin, View):
    template_name = 'expenses/goal_detail.html'

    def get(self, request, pk):
        goal = get_object_or_404(SavingsGoal, pk=pk, user=request.user)
        contributions = goal.contributions.all().order_by('-date', '-created_at')
        form = GoalContributionForm()
        
        # Check if goal is locked
        profile = request.user.profile
        is_locked = False
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 1
            goals_ids = list(SavingsGoal.objects.filter(user=request.user).order_by('created_at').values_list('id', flat=True))
            if goal.id in goals_ids and goals_ids.index(goal.id) >= limit:
                 is_locked = True
        
        context = {
            'goal': goal,
            'is_locked': is_locked,
            'contributions': contributions,
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        goal = get_object_or_404(SavingsGoal, pk=pk, user=request.user)
        
        # Check if goal is locked due to subscription limits
        profile = request.user.profile
        if not profile.is_pro:
            limit = 3 if profile.is_plus else 1
            goals_ids = list(SavingsGoal.objects.filter(user=request.user).order_by('created_at').values_list('id', flat=True))
            if goal.id in goals_ids and goals_ids.index(goal.id) >= limit:
                 messages.error(request, _("This savings goal is locked because it exceeds your plan's limit. Upgrade to unlock it!"))
                 return redirect('goal-detail', pk=goal.pk)

        # Handle JS fetch to clear confetti
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                if data.get('clear_confetti'):
                    if 'trigger_confetti' in request.session:
                        del request.session['trigger_confetti']
                    return JsonResponse({'success': True})
            except:
                pass

        contributions = goal.contributions.all().order_by('-date', '-created_at')
        form = GoalContributionForm(request.POST)

        if form.is_valid():
            contribution = form.save(commit=False)
            contribution.goal = goal
            contribution.save() # This triggers the save() method in models which updates the goal's current_amount
            
            messages.success(request, _("Funds added successfully!"))
            request.session['trigger_confetti'] = True
            
            return redirect('goal-detail', pk=goal.pk)

        context = {
            'goal': goal,
            'contributions': contributions,
            'form': form
        }
        return render(request, self.template_name, context)


class YearInReviewView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/year_in_review.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.profile.is_plus:
            messages.info(request, "Year in Review is a Premium feature. Upgrade to Plus or Pro to unlock your personalized financial story!")
            return redirect('pricing')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get requested year, default to current year
        requested_year = self.kwargs.get('year')
        current_year = datetime.now().year
        
        if not requested_year:
            # If it's Jan or Feb, they probably want to see last year's review
            if datetime.datetime.now().month <= 2:
                requested_year = current_year - 1
            else:
                requested_year = current_year
                
        context['year'] = requested_year
        context['review_data'] = generate_year_in_review_data(self.request.user, requested_year)
        
        return context
